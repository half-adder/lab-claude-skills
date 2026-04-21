# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
"""
workbook_helpers.py -- shared primitives for protocol-planner workbooks.

Importable as a module by uv-script users:

    # /// script
    # dependencies = ["openpyxl"]
    # ///
    import sys
    sys.path.insert(0, "/Users/sean/code/lab-claude-skills/protocol-planner/scripts")
    from workbook_helpers import (
        put, write_stocks_table, write_unit_factor_table,
        buffer_block, per_sample_mix_block, bench_add_block,
        save_with_full_calc, STYLES,
    )

All functions take an openpyxl Worksheet `ws` as their first arg (except where noted).
"""
from __future__ import annotations

from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.worksheet import Worksheet


# --- Shared styles -----------------------------------------------------------

@dataclass(frozen=True)
class _Styles:
    H1: Font = Font(bold=True, size=14)
    H2: Font = Font(bold=True, size=12)
    BOLD: Font = Font(bold=True)
    ITAL: Font = Font(italic=True, color="555555")
    HDR_FILL: PatternFill = PatternFill("solid", fgColor="E8E8E8")
    BLOCK_FILL: PatternFill = PatternFill("solid", fgColor="F3F7FF")


STYLES = _Styles()


# --- Cell write helper -------------------------------------------------------

def put(
    ws: Worksheet,
    cell: str,
    value,
    font: Font | None = None,
    fill: PatternFill | None = None,
    number_format: str | None = None,
) -> None:
    """Write a value (or formula) to a cell with optional styling.

    Formula strings must start with '='.  Plain text labels must NOT start with
    '='; Excel will try to evaluate them and emit #NAME?.
    """
    c = ws[cell]
    c.value = value
    if font is not None:
        c.font = font
    if fill is not None:
        c.fill = fill
    if number_format is not None:
        c.number_format = number_format


# --- Unit-factor table -------------------------------------------------------

DEFAULT_UNITS = [
    ("M", 1),
    ("mM", 0.001),
    ("µM", 0.000001),
    ("nM", 0.000000001),
    ("%", 1),
    ("mg/mL", 1),
    ("µg/mL", 0.001),
]


def write_unit_factor_table(
    ws: Worksheet,
    anchor_cell: str = "F7",
    units: list[tuple[str, float]] | None = None,
) -> str:
    """Write a Unit | Factor lookup table starting at anchor_cell.

    Returns the absolute range string (e.g. "$F$8:$G$14") for use in VLOOKUP.
    """
    units = units or DEFAULT_UNITS
    col = anchor_cell[0]
    row = int(anchor_cell[1:])
    next_col = chr(ord(col) + 1)
    put(ws, f"{col}{row}", "Unit", STYLES.BOLD, STYLES.HDR_FILL)
    put(ws, f"{next_col}{row}", "Factor", STYLES.BOLD, STYLES.HDR_FILL)
    for i, (u, f) in enumerate(units):
        r = row + 1 + i
        put(ws, f"{col}{r}", u)
        put(ws, f"{next_col}{r}", f)
    last_row = row + len(units)
    return f"${col}${row+1}:${next_col}${last_row}"


def unit_factor_expr(unit_cell: str, units_range: str) -> str:
    """Return a VLOOKUP-with-IFERROR expression that resolves a unit string to
    its numeric factor.  Falls back to 1 for unknown units.
    """
    return f"IFERROR(VLOOKUP({unit_cell},{units_range},2,FALSE),1)"


# --- Stocks table ------------------------------------------------------------

def write_stocks_table(
    ws: Worksheet,
    start_row: int,
    stocks: list[tuple[str, float, str, str]],
) -> dict[str, int]:
    """Write the Stock display | Conc | Unit | Chemical table.

    `stocks` is a list of (key, conc, unit, chemical).  Returns a dict mapping
    each key to its row number so callers can reference the stock by cell.

    The display column (A) is a formula that rebuilds when Conc/Unit/Chemical
    change.  '%' has no space before the symbol; other units do.
    """
    put(ws, f"A{start_row-1}", "Stock (display)", STYLES.BOLD, STYLES.HDR_FILL)
    put(ws, f"B{start_row-1}", "Conc", STYLES.BOLD, STYLES.HDR_FILL)
    put(ws, f"C{start_row-1}", "Unit", STYLES.BOLD, STYLES.HDR_FILL)
    put(ws, f"D{start_row-1}", "Chemical", STYLES.BOLD, STYLES.HDR_FILL)

    stock_row: dict[str, int] = {}
    for i, (key, conc, unit, chem) in enumerate(stocks):
        r = start_row + i
        stock_row[key] = r
        put(
            ws,
            f"A{r}",
            f'=IF(C{r}="%",B{r}&"% "&D{r},B{r}&" "&C{r}&" "&D{r})',
        )
        put(ws, f"B{r}", conc)
        put(ws, f"C{r}", unit)
        put(ws, f"D{r}", chem)
    return stock_row


# --- Buffer block -----------------------------------------------------------

def buffer_block(
    ws: Worksheet,
    start_row: int,
    title: str,
    per_ip_ul: float,
    per_ip_note: str,
    components: list[tuple[str, float, str]],  # (stock_key, final_val, final_unit)
    stock_row: dict[str, int],
    units_range: str,
    ip_cell: str = "$B$4",
    deadvol_cell: str = "$B$3",
    balance_label: str | None = "ddH2O (balance)",
) -> int:
    """Write one buffer block. Returns the next free row (with one blank row gap)."""
    r = start_row
    put(ws, f"A{r}", title, STYLES.H2, STYLES.BLOCK_FILL)
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    put(ws, f"A{r}", "Per-IP volume", STYLES.BOLD)
    put(ws, f"B{r}", per_ip_ul)
    put(ws, f"C{r}", "µL")
    put(ws, f"D{r}", per_ip_note, STYLES.ITAL)
    per_ip_cell = f"B{r}"
    r += 1
    put(ws, f"A{r}", "Total to prepare", STYLES.BOLD)
    put(ws, f"B{r}", f"={per_ip_cell}*{ip_cell}*{deadvol_cell}/1000", number_format="0.00")
    put(ws, f"C{r}", "mL")
    put(ws, f"D{r}", f"={per_ip_cell}*{ip_cell}*{deadvol_cell}", number_format="0")
    put(ws, f"E{r}", "µL")
    total_ul_cell = f"D{r}"
    r += 2

    put(ws, f"A{r}", "Component", STYLES.BOLD, STYLES.HDR_FILL)
    put(ws, f"B{r}", "Final val", STYLES.BOLD, STYLES.HDR_FILL)
    put(ws, f"C{r}", "Unit", STYLES.BOLD, STYLES.HDR_FILL)
    put(ws, f"D{r}", "Volume (µL)", STYLES.BOLD, STYLES.HDR_FILL)
    put(ws, f"E{r}", "(stock)", STYLES.BOLD, STYLES.HDR_FILL)
    r += 1

    first_vol_row = r
    for stock_key, final_val, final_unit in components:
        srow = stock_row[stock_key]
        put(ws, f"A{r}", f"=$A${srow}")
        put(ws, f"B{r}", final_val)
        put(ws, f"C{r}", final_unit)
        f_final = unit_factor_expr(f"C{r}", units_range)
        f_stock = unit_factor_expr(f"$C${srow}", units_range)
        put(
            ws,
            f"D{r}",
            f"={total_ul_cell}*(B{r}*{f_final})/($B${srow}*{f_stock})",
            number_format="0.00",
        )
        put(ws, f"E{r}", f"=$A${srow}", STYLES.ITAL)
        r += 1
    last_vol_row = r - 1

    if balance_label is not None:
        put(ws, f"A{r}", balance_label)
        put(
            ws,
            f"D{r}",
            f"={total_ul_cell}-SUM(D{first_vol_row}:D{last_vol_row})",
            number_format="0.00",
        )
        r += 1

    put(ws, f"A{r}", "TOTAL (sanity check)", STYLES.ITAL)
    put(
        ws,
        f"D{r}",
        f"=SUM(D{first_vol_row}:D{r-1})",
        STYLES.ITAL,
        number_format="0.00",
    )
    put(ws, f"E{r}", f"should equal {total_ul_cell}", STYLES.ITAL)
    return r + 2


# --- Per-sample mix block (no stocks, no water balance) ---------------------

def per_sample_mix_block(
    ws: Worksheet,
    start_row: int,
    title: str,
    per_sample_components: list[tuple[str, float, str]],  # (name, µL/sample, note)
    ip_cell: str = "$B$4",
    deadvol_cell: str = "$B$3",
) -> int:
    """Write a per-sample mastermix block (no stocks-table lookup, no water balance).

    Returns next free row.
    """
    total_per_ip = sum(c[1] for c in per_sample_components)
    r = start_row
    put(ws, f"A{r}", title, STYLES.H2, STYLES.BLOCK_FILL)
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    put(ws, f"A{r}", "Per-IP volume", STYLES.BOLD)
    put(ws, f"B{r}", total_per_ip)
    put(ws, f"C{r}", "µL")
    put(ws, f"D{r}", " + ".join(str(c[1]) for c in per_sample_components), STYLES.ITAL)
    per_ip_cell = f"B{r}"
    r += 1
    put(ws, f"A{r}", "Total to prepare", STYLES.BOLD)
    put(ws, f"B{r}", f"={per_ip_cell}*{ip_cell}*{deadvol_cell}/1000", number_format="0.00")
    put(ws, f"C{r}", "mL")
    put(ws, f"D{r}", f"={per_ip_cell}*{ip_cell}*{deadvol_cell}", number_format="0")
    put(ws, f"E{r}", "µL")
    r += 2
    put(ws, f"A{r}", "Component", STYLES.BOLD, STYLES.HDR_FILL)
    put(ws, f"B{r}", "µL / sample", STYLES.BOLD, STYLES.HDR_FILL)
    put(ws, f"C{r}", "Total µL", STYLES.BOLD, STYLES.HDR_FILL)
    put(ws, f"D{r}", "Notes", STYLES.BOLD, STYLES.HDR_FILL)
    r += 1
    for name, per_sample, note in per_sample_components:
        put(ws, f"A{r}", name)
        put(ws, f"B{r}", per_sample)
        put(
            ws,
            f"C{r}",
            f"={per_sample}*{ip_cell}*{deadvol_cell}",
            number_format="0.00",
        )
        put(ws, f"D{r}", note, STYLES.ITAL)
        r += 1
    return r + 1


# --- Bench-add reagent block (added at bench, not premixed) ------------------

def bench_add_block(
    ws: Worksheet,
    start_row: int,
    title: str,
    per_ip_ul: float,
    stock_display_cell: str,
    note_suffix: str = " directly into each tube",
    ip_cell: str = "$B$4",
    deadvol_cell: str = "$B$3",
) -> int:
    """Write a small block for a reagent added separately at the bench.

    `stock_display_cell` is an absolute reference to the stock's display cell
    (e.g. "$A$15") so the note rebuilds when the stock changes.
    """
    r = start_row
    put(ws, f"A{r}", title, STYLES.H2, STYLES.BLOCK_FILL)
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    put(ws, f"A{r}", "Per-IP volume", STYLES.BOLD)
    put(ws, f"B{r}", per_ip_ul)
    put(ws, f"C{r}", "µL")
    put(ws, f"D{r}", f'={stock_display_cell}&"{note_suffix}"', STYLES.ITAL)
    r += 1
    put(ws, f"A{r}", "Total to aliquot", STYLES.BOLD)
    put(
        ws,
        f"B{r}",
        f"={per_ip_ul}*{ip_cell}*{deadvol_cell}",
        number_format="0.00",
    )
    put(ws, f"C{r}", f'="µL of "&{stock_display_cell}&" stock"')
    return r + 2


# --- Save -------------------------------------------------------------------

def save_with_full_calc(wb: Workbook, path: str) -> None:
    """Save workbook with fullCalcOnLoad=True so Excel recomputes on open.

    Without this, formulas can render blank in Excel until the user clicks
    each cell and presses Enter.
    """
    wb.calculation = CalcProperties(fullCalcOnLoad=True)
    wb.save(path)
